# -*- coding:utf-8 -*-
"""
author: byangg
datettime: 8/5/2020 16:17
"""

'''
更新买卖点记录中的收益计算
'''


from openpyxl import load_workbook
import pandas as pd
from get_data_tspro import get_daily_basic_by_date, engine
import shutil

excel_path = r"D:\synchronize\投资理财\买卖点模拟&记录.xlsx" #买卖点模拟&记录
bk_path = r"D:\synchronize\投资理财\买卖点模拟&记录_bk.xlsx"
shutil.copy(excel_path, bk_path)

wb = load_workbook(excel_path)
ws = wb['收益计算']

## 更新名称及数量
row_n = 2
while True:
    if ws[f'b{row_n}'].value is None:
        break
    ws[f'w{row_n}'] = ws[f'b{row_n}'].value
    ws[f'x{row_n}'] = -ws[f'e{row_n}'].value * ws[f'f{row_n}'].value if abs(ws[f'e{row_n}'].value) > 10 else 0
    row_n += 1


## 获取持有股票名称
row_n = 2
indices = []
names = []
while True:
    value = ws[f'o{row_n}'].value
    if value == "总计":
        break
    indices.append(row_n)
    names.append(value)
    row_n += 1

ts_codes = pd.read_sql(f'''select * from stock_basic where name in ('{"','".join(names)}')''', engine)
daily_basic_df = get_daily_basic_by_date().reset_index(col_fill='ts_code')
daily_basic_df['total_mv'] = daily_basic_df['total_mv'] / 10000
df1 = pd.merge(ts_codes, daily_basic_df, on='ts_code').loc[:,['name','close','total_mv']]

# total_ = 0
for _, row in df1.iterrows():
    i = names.index(row['name'])
    idx = indices[i]
    ws[f'Q{idx}'] = row['close']
    # ws[f'R{idx}'] = ws[f'P{idx}'].value * row['close']
    ws[f'S{idx}'] = row['total_mv']
    # total_ += ws[f'R{idx}'].value
# ws[f'R{idx+1}'] = total_

wb.save(excel_path)


print('done')